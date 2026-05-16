"""
Client Chart Generator — pulls Stripe invoice data and exports to XLSX.

Auto-filled from Stripe: Client, Email, Value, Average Retainer, Join Date, Churn Date
Manual columns (left blank): Churn Risk, Priority, SKAG, Total Retainer Clients
"""

import os
import stripe
from datetime import datetime, timezone
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

stripe.api_key = os.getenv("STRIPE_SECRET_KEY")
if not stripe.api_key:
    raise ValueError("STRIPE_SECRET_KEY not set in .env")

COLUMNS = [
    "Client",
    "Email",
    "Value ($/mo)",
    "Average Retainer ($/mo)",
    "Join Date",
    "Churn Date",
    "Churn Risk",
    "Priority",
    "SKAG",
    "Total Retainer Clients",
]

# Clients with no paid invoice in this many days are considered churned
CHURN_THRESHOLD_DAYS = 60

HEADER_COLOR = "1A1A2E"
HEADER_FONT_COLOR = "FFFFFF"
ACTIVE_ROW_COLOR = "EAF4FB"
CHURNED_ROW_COLOR = "FDF0F0"
ALT_ROW_COLOR = "F7F7F7"


def fetch_clients():
    now = datetime.now(timezone.utc).timestamp()
    churn_cutoff = now - (CHURN_THRESHOLD_DAYS * 86400)

    print("Fetching all Stripe customers...")
    all_customers = list(stripe.Customer.list(limit=100).auto_paging_iter())
    print(f"  {len(all_customers)} customers found. Fetching invoices...")

    rows = []

    for i, customer in enumerate(all_customers):
        cid = customer.id

        # Get all paid invoices for this customer
        paid_invoices = list(
            stripe.Invoice.list(
                customer=cid,
                status="paid",
                limit=100,
            ).auto_paging_iter()
        )

        if not paid_invoices:
            continue  # skip customers with no paid invoices

        amounts = [inv.amount_paid / 100 for inv in paid_invoices if inv.amount_paid > 0]
        if not amounts:
            continue

        # Sort invoices by date
        paid_invoices.sort(key=lambda inv: inv.created)

        latest_invoice = paid_invoices[-1]
        first_invoice = paid_invoices[0]

        latest_amount = latest_invoice.amount_paid / 100
        avg_amount = round(sum(amounts) / len(amounts), 2)

        join_date = datetime.fromtimestamp(first_invoice.created).strftime("%Y-%m-%d")

        # Churned = no paid invoice in the last CHURN_THRESHOLD_DAYS days
        is_churned = latest_invoice.created < churn_cutoff
        churn_date = ""
        if is_churned:
            churn_date = datetime.fromtimestamp(latest_invoice.created).strftime("%Y-%m-%d")

        rows.append({
            "Client": customer.name or "",
            "Email": customer.email or "",
            "Value ($/mo)": round(latest_amount, 2),
            "Average Retainer ($/mo)": avg_amount,
            "Join Date": join_date,
            "Churn Date": churn_date,
            "Churn Risk": "",
            "Priority": "",
            "SKAG": "",
            "Total Retainer Clients": "",
            "_churned": is_churned,
        })

        if (i + 1) % 20 == 0:
            print(f"  Processed {i + 1}/{len(all_customers)} customers...")

    # Sort: active clients by value desc, then churned by churn date desc
    active = sorted([r for r in rows if not r["_churned"]], key=lambda r: r["Value ($/mo)"], reverse=True)
    churned = sorted([r for r in rows if r["_churned"]], key=lambda r: r["Churn Date"], reverse=True)

    print(f"\nActive clients: {len(active)}")
    print(f"Churned clients: {len(churned)}")

    return active + churned


def build_xlsx(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clients"
    ws.freeze_panes = "A2"

    # Header
    header_fill = PatternFill("solid", fgColor=HEADER_COLOR)
    header_font = Font(bold=True, color=HEADER_FONT_COLOR, size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    ws.row_dimensions[1].height = 36

    # Data rows
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, row in enumerate(rows, start=2):
        is_churned = row.pop("_churned")
        is_alt = (row_idx % 2 == 0)

        if is_churned:
            fill_color = CHURNED_ROW_COLOR
        elif is_alt:
            fill_color = ALT_ROW_COLOR
        else:
            fill_color = ACTIVE_ROW_COLOR

        fill = PatternFill("solid", fgColor=fill_color)

        for col_idx, col_name in enumerate(COLUMNS, start=1):
            value = row.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(
                horizontal="left" if col_idx <= 2 else "center",
                vertical="center"
            )

        ws.row_dimensions[row_idx].height = 22

    # Apply border to header too
    for col_idx in range(1, len(COLUMNS) + 1):
        ws.cell(row=1, column=col_idx).border = border

    # Column widths
    widths = {
        "Client": 28,
        "Email": 30,
        "Value ($/mo)": 16,
        "Average Retainer ($/mo)": 22,
        "Join Date": 14,
        "Churn Date": 14,
        "Churn Risk": 14,
        "Priority": 12,
        "SKAG": 10,
        "Total Retainer Clients": 22,
    }
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 16)

    filename = f"client-chart-{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    output_path = os.path.join(os.path.dirname(__file__), filename)
    wb.save(output_path)
    print(f"\nSaved: {output_path}")
    return output_path


if __name__ == "__main__":
    rows = fetch_clients()
    if not rows:
        print("No clients with paid invoices found.")
    else:
        build_xlsx(rows)
